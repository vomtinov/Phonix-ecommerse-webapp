from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
import logging
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
from datetime import datetime
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from authentication.models import Category,Product,Brand,Variant,ProductImage,Order,Offer,Coupon,CouponUsage
from .forms import ProductForm,VariantForm,VariantImageFormSet,OfferForm
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.http import JsonResponse  # Add this import
from django.contrib.auth.decorators import user_passes_test
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_POST
from django.db import transaction
from django.views.decorators.cache import never_cache
from django.utils import timezone
from django.db.models import Sum, Count, Q,Avg
from datetime import timedelta
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth, TruncYear
from reportlab.lib.styles import ParagraphStyle
import openpyxl



User = get_user_model()  # Get the custom user model
 
#this view for admin login
def admin_login(request):
    if request.method == "POST":
        email = request.POST.get("email")  
        password = request.POST.get("password")
        
        if not email or not password:
            messages.error(request, "Email and password are required")
            return render(request, "admin_login.html")

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            messages.error(request, "Invalid user")
            return render(request, "admin_login.html")

        if user.is_superuser and not user.is_staff:
            messages.error(request, "Invalid user")
            return render(request, "admin_login.html")

        user = authenticate(request, username=email, password=password)  

        if user is None:
            messages.error(request, "Invalid email or password")  
            return render(request, "admin_login.html")

        if not user.is_staff:
            messages.error(request, "You do not have admin access")
            return render(request, "admin_login.html")

        login(request, user)
        return redirect("admin_panel")

    return render(request, "admin_login.html")

#this view for admin panel
def admin_panel(request):
    if not request.user.is_staff:
        return redirect("admin_login")
    return render(request, "admin_panel.html")

#this view for admin dashboard
def dashboard_view(request):
    if not request.user.is_staff:
        return redirect("admin_login")
    return render(request, "dashboard.html")

#this view for displaying all users 
def users_view(request):
    if not request.user.is_staff:
        return redirect("admin_login")

    search_query = request.GET.get("search", "")
    clear = request.GET.get("clear", "no")
    if clear == "yes" or not search_query:
        search_query = ""

    users = User.objects.filter(username__icontains=search_query, is_staff=False).order_by("-id")

    paginator = Paginator(users, 5)  
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "users.html", {"users": page_obj})

#this view for blocking the user
def block_user(request, user_id):
    if not request.user.is_staff:
        return redirect("admin_login")

    user = get_object_or_404(User, id=user_id)
    user.is_active = False
    user.save()
    
    messages.success(request, f"{user.username} has been blocked successfully!")
    return redirect("users")

#this view for un-blocking the user
def unblock_user(request, user_id):
    if not request.user.is_staff:
        return redirect("admin_login")

    user = get_object_or_404(User, id=user_id)
    user.is_active = True
    user.save()

    messages.success(request, f"{user.username} has been unblocked successfully!")
    return redirect("users")         

#this view for category list
def category_list(request):
    if not request.user.is_staff:
        return redirect("admin_login")

    search_query = request.GET.get("search", "")
    clear = request.GET.get("clear", "no")  

    if clear == "yes" or not search_query:
        search_query = ""

    categories = Category.objects.filter(name__icontains=search_query, is_deleted=False).order_by("-created_at" )

    paginator = Paginator(categories, 5)  
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "category.html", {"categories": page_obj, "request": request})

#this view for add-category list
def add_category(request):
    if not request.user.is_staff:
        return redirect("admin_login")

    if request.method == "POST":
        name = request.POST.get("name")
        description = request.POST.get("description")
        status = request.POST.get("status", "1") 

        # Server-side validation
        errors = []
        
        # Validate name
        if not name:
            errors.append("Category name is required")
        elif len(name) > 255:
            errors.append("Category name cannot exceed 255 characters")
        
        # Check if name contains only numbers
        if name and name.strip().isdigit():
            errors.append("Category name cannot contain only numbers")
            
        # Validate description  
        if not description:
            errors.append("Description is required")
            
        # Check for special characters in name (optional)
        import re
        if name and re.search(r'[!@#$%^&*()_+\=\[\]{};:"\\|,.<>\/?]+', name):
            errors.append("Category name should not contain special characters")
            
        # Check uniqueness
        if name and Category.objects.filter(name__iexact=name, is_deleted=False).exists():
            errors.append("Category with this name already exists")
            
        # If any validation errors, return them
        if errors:
            for error in errors:
                messages.error(request, error)
            return redirect("add_category")

        status_value = True if status == "1" else False

        try:
            Category.objects.create(
                name=name, 
                description=description, 
                status=status_value, 
                is_deleted=False
            )
            messages.success(request, "Category added successfully!")
            return redirect("category_list")  
        except Exception as e:
            messages.error(request, f"An error occurred while adding the category: {str(e)}")
            return redirect("add_category")

    return render(request, "add_category.html")


#this view for edit-category list
def edit_category(request, category_id):
    if not request.user.is_staff:
        return redirect("admin_login")

    category = get_object_or_404(Category, id=category_id, is_deleted=False)
    
    if request.method == "POST":
        name = request.POST.get("name")
        description = request.POST.get("description")
        status = 'status' in request.POST  

        if not name or not description:
            messages.error(request, "All fields are required!")
            return render(request, "edit_category.html", {"category": category})

        if Category.objects.filter(name=name, is_deleted=False).exclude(id=category_id).exists():
            messages.error(request, "Category name already exists!")
            return render(request, "edit_category.html", {"category": category})

        category.name = name
        category.description = description
        category.status = status
        category.save()

        messages.success(request, "Category updated successfully!")
        return redirect("category_list")

    return render(request, "edit_category.html", {"category": category})

#this view for delete-category list
def delete_category(request, category_id):
    if not request.user.is_staff:
        return redirect("admin_login")

    category = get_object_or_404(Category, id=category_id)
    confirm = request.GET.get('confirm', 'no')  

    if confirm == 'yes':
        category.is_deleted = True
        category.save()
        messages.success(request, "Category soft-deleted successfully!")
        return redirect("category_list")
    else:
        
        return render(request, "confirm_action.html", {
            "category": category,
            "action": "delete",
            "object_type": "Category",
            "return_url": "category_list",
            "next_url": f"/delete_category/{category_id}/?confirm=yes"
        })

#this view for brand list
def brand_list(request):
    if not request.user.is_staff:
        return redirect("admin_login")
    
    # Get the queryset - make sure to execute it with ()
    brand_list = Brand.objects.all().order_by('-id')
    print(f"Type of brand_list: {type(brand_list)}")  # Should print <class 'django.db.models.query.QuerySet'>
    print(f"Brand count: {brand_list.count()}")  # Should print number of brands
    
    # Set up pagination - 10 items per page
    paginator = Paginator(brand_list, 10)
    print(f"Paginator created with {paginator.count} items")  # Should match brand count
    
    page = request.GET.get('page')
    try:
        brands = paginator.page(page)
        print(f"Current page: {brands.number}, Items: {len(brands)}")
    except PageNotAnInteger:
        brands = paginator.page(1)
        print("Defaulting to page 1")
    except EmptyPage:
        brands = paginator.page(paginator.num_pages)
        print(f"Defaulting to last page: {paginator.num_pages}")
    
    return render(request, 'brand_list.html', {'brands': brands})

#this view for add brand list
def add_brand(request):
    if not request.user.is_staff:
        return redirect("admin_login")
    if request.method == "POST":
        name = request.POST.get('name')
        description = request.POST.get('description')
        status = request.POST.get('status') == '1'  

        Brand.objects.create(name=name, description=description, status=status)
        return redirect('brand_list')

    return render(request, 'add_brand.html')


#this view for edit-brand list
def edit_brand(request, brand_id):
    if not request.user.is_staff:
        return redirect("admin_login")
    
    brand = get_object_or_404(Brand, id=brand_id)

    if request.method == "POST":
        brand.name = request.POST.get('name')
        brand.description = request.POST.get('description')
        new_status = request.POST.get('status') == '1'
        old_status = brand.status
        
        if not new_status:
            Product.objects.filter(brand=brand).update(status=False)
        elif new_status and not old_status:
            Product.objects.filter(brand=brand).update(status=True)
            
        brand.status = new_status
        brand.save()
        return redirect('brand_list')
    
    return render(request, 'edit_brand.html', {'brand': brand})


#this view for product list
def product_list(request):
    if not request.user.is_staff:
        return redirect("admin_login")
    
    products = Product.objects.filter(is_deleted=False).prefetch_related('variants__images')
    return render(request, 'product_list.html', {'products': products})

#this view for adding new product 
def add_product(request):
    if not request.user.is_staff:
        return redirect("admin_login")
    
    categories = Category.objects.all()
    brands = Brand.objects.all()

    if request.method == "POST":
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save()
            rams = request.POST.getlist('ram[]')
            storages = request.POST.getlist('storage[]')
            colors = request.POST.getlist('color[]')
            prices = request.POST.getlist('price[]')
            stocks = request.POST.getlist('stock[]')
            if not rams:
                product.delete()
                messages.error(request, "At least one variant is required.")
                return render(request, 'add_product.html', {
                    'form': form,
                    'categories': categories,
                    'brands': brands,
                })
            # Process each variant
            for i in range(len(rams)):
                variant_data = {
                    'ram': rams[i],
                    'storage': storages[i],
                    'color': colors[i],
                    'price': prices[i],
                    'stock': stocks[i],
                }
                variant_form = VariantForm(variant_data)
                if variant_form.is_valid():
                    variant = variant_form.save(commit=False)
                    variant.product = product
                    variant.save()
                    # Handle images for this variant
                    image_files = request.FILES.getlist(f'variant_images_{i+1}[]')
                    if len(image_files) < 3 or len(image_files) > 6:
                        variant.delete()
                        product.delete()
                        messages.error(request, f"Variant {i+1} must have between 3 and 6 images.")
                        return render(request, 'add_product.html', {
                            'form': form,
                            'categories': categories,
                            'brands': brands,
                        })
                    for image_file in image_files:
                        ProductImage.objects.create(
                            variant=variant,
                            image=image_file
                        )
                else:
                    product.delete()
                    messages.error(request, f"Errors in Variant {i+1}: {variant_form.errors}")
                    return render(request, 'add_product.html', {
                        'form': form,
                        'categories': categories,
                        'brands': brands,
                    })
            messages.success(request, "Product and variants added successfully!")
            return redirect('product_list')
        else:
            messages.error(request, "Please correct the errors in the product details.")
    else:
        form = ProductForm()
    context = {
        'form': form,
        'categories': categories,
        'brands': brands,
    }
    return render(request, 'add_product.html', context)

#this view for edit-product list
def edit_product(request, product_id):
    if not request.user.is_staff:
        return redirect("admin_login")
    
    product = get_object_or_404(Product, id=product_id, is_deleted=False)

    categories = Category.objects.all()
    brands = Brand.objects.all()
    variants = product.variants.all()

    if request.method == "POST":
        product.name = request.POST.get("name")
        product.description = request.POST.get("description")
        product.category_id = request.POST.get("category")
        product.brand_id = request.POST.get("brand")
        product.save()

        rams = request.POST.getlist('ram[]')
        storages = request.POST.getlist('storage[]')
        colors = request.POST.getlist('color[]')
        prices = request.POST.getlist('price[]')
        stocks = request.POST.getlist('stock[]')
        variant_ids = request.POST.getlist('variant_id[]')

        for i in range(len(rams)):
            variant_data = {
                'ram': rams[i],
                'storage': storages[i],
                'color': colors[i],
                'price': prices[i],
                'stock': stocks[i],
            }
            if i < len(variant_ids) and variant_ids[i]:  # Existing variant
                variant = Variant.objects.get(id=variant_ids[i], product=product)
                for key, value in variant_data.items():
                    setattr(variant, key, value)
                variant.save()
                # Handle images for existing variant
                image_files = request.FILES.getlist(f'variant_images_{i+1}[]')
                for image_file in image_files:
                    img = ProductImage(variant=variant, image=image_file)
                    img.save()
            else:  # New variant
                variant = Variant(product=product, **variant_data)
                variant.save()
                image_files = request.FILES.getlist(f'variant_images_{i+1}[]')
                for image_file in image_files:
                    img = ProductImage(variant=variant, image=image_file)
                    img.save()

        removed_images = request.POST.get("removed_images", "").split(",")
        for image_id in removed_images:
            if image_id.isdigit():
                ProductImage.objects.filter(id=int(image_id)).delete()

        messages.success(request, "Product updated successfully!")
        return redirect('product_list')

    context = {
        'product': product,
        'categories': categories,
        'brands': brands,
        'variants': variants,
    }
    return render(request, 'edit_product.html', context)

#this view for adding new varient for the product
@csrf_exempt
def add_variant(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        form = VariantForm(request.POST,request.FILES)
        if form.is_valid():
            variant = form.save(commit=False)
            variant.product = product
            variant.save()
            return JsonResponse({"success": True, "message": "Variant added successfully!"})
    
        else:
            print("Form errors:", form.errors)
            return JsonResponse({"success": False, "error": form.errors})
    
    return render(request, 'add_variant.html', {'product': product})

logger = logging.getLogger(__name__)

#this view for deleteting product 
@require_POST
def delete_product(request, product_id):
    if not request.user.is_staff:
        return redirect("admin_login")
    try:
        # Log the incoming request details
        logger.info(f"Delete product request received for product {product_id}")
        logger.info(f"Request method: {request.method}")
        logger.info(f"Request body: {request.body}")

        product = get_object_or_404(Product, id=product_id)
        
        product.is_deleted = True
        product.status = False
        product.save()
        
        logger.info(f"Product {product_id} soft deleted")
        return JsonResponse({
            "success": True, 
            "message": f"Product '{product.name}' deleted successfully!"
        })
    
    except Exception as e:
        logger.error(f"Error in delete_product: {str(e)}")
        return JsonResponse({
            "success": False, 
            "error": str(e)
        }, status=500)
    
import json

#this view for toggle-status list
@require_POST
def toggle_product_status(request, product_id):
    if not request.user.is_staff:
        return redirect("admin_login")
    try:
        # Log the incoming request details
        logger.info(f"Toggle status request received for product {product_id}")
        logger.info(f"Request method: {request.method}")
        logger.info(f"Request body: {request.body}")
        
        # Try to parse the request body
        try:
            data = json.loads(request.body)
            logger.info(f"Parsed data: {data}")
        except json.JSONDecodeError:
            logger.error("Failed to parse JSON from request body")
            return JsonResponse({"success": False, "error": "Invalid JSON"}, status=400)

        product = get_object_or_404(Product, id=product_id)
        
        # Toggle status
        product.status = not product.status
        product.save()
        
        logger.info(f"Product {product_id} status updated to {product.status}")
        return JsonResponse({
            "success": True, 
            "status": product.status
        })
    
    except Exception as e:
        logger.error(f"Error in toggle_product_status: {str(e)}")
        return JsonResponse({
            "success": False, 
            "error": str(e)
        }, status=500)

#this view for seeing all orders list
def order_list(request):
    if not request.user.is_staff:
        return redirect("admin_login")
    orders_queryset = Order.objects.all().order_by('-created_at')
    raw_statuses = Order.objects.values('id', 'status')
    
    paginator = Paginator(orders_queryset, 5)
    page = request.GET.get('page', 1)
    
    try:
        orders = paginator.page(page)
    except PageNotAnInteger:
        orders = paginator.page(1)
    except EmptyPage:
        orders = paginator.page(paginator.num_pages)
    
    return render(request, 'order.html', {'orders': orders})

#this view for changing the order status
@csrf_exempt  
def update_order_status(request):
    if not request.user.is_staff:
        return redirect("admin_login")
    
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Invalid request method"})
    
    order_id = request.POST.get("order_id")
    status = request.POST.get("status")
    
    try:
        order = Order.objects.get(id=order_id)
        old_status = order.status
        if old_status != status:
            order.status = status  
            if status == "Cancelled" and old_status in ["Pending", "Confirmed", "Shipped"]:
                for item in order.items.all():
                    product = item.product
                    product.stock += item.quantity
                    product.quantity = product.stock
                    product.save()
            order.save()  
            updated_order = Order.objects.get(id=order_id)         
            if updated_order.status != status:
                print(f"WARNING: Status mismatch - Expected {status}, Got {updated_order.status}")
        return JsonResponse({"success": True, "status": updated_order.status})
    except Order.DoesNotExist:
        return JsonResponse({"success": False, "error": "Order not found"})
    except Exception as e:
        print(f"Error updating order {order_id}: {str(e)}")
        return JsonResponse({"success": False, "error": str(e)})

#this view for approval for returning
@csrf_protect
@require_POST
def approve_return(request, order_id):
    if not request.user.is_staff:
        return redirect("admin_login")
    try:
        order = Order.objects.get(id=order_id)
        
        if order.status != "Return Requested":
            return JsonResponse({"success": False, "message": "Order is not in Return Requested status"})
        
        print(f"Approving return for Order {order_id} - Current status: {order.status}")
        order.status = "Returned"
        order.save()
        
        for item in order.items.all():
            product = item.product
            product.stock += item.quantity
            product.quantity = product.stock  
            product.save()
        
        updated_order = Order.objects.get(id=order_id)
        
        return JsonResponse({
            "success": True, 
            "message": "Return approved", 
            "status": updated_order.status
        })
    except Order.DoesNotExist:
        return JsonResponse({"success": False, "message": "Order not found"})
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)})
    
#this view for seeing all orders details
def order_detail(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    return render(request, 'order_fulldetail.html', {'order': order})

#this view for updating stock for the varients
def get_product_quantities(request):
    if not request.user.is_staff:
        return redirect("admin_login")

    variants = Variant.objects.filter(product__is_deleted=False).values('product__id', 'stock')
    # Aggregate(sum) stock by product ID 
    stocks = {}
    for variant in variants:
        product_id = str(variant['product__id'])
        stock = variant['stock']
        if product_id in stocks:
            stocks[product_id] += stock  
        else:
            stocks[product_id] = stock
    return JsonResponse({'stocks': stocks})  

#this view for seeing all offer details
def offer_management(request):
    if not request.user.is_staff:
        return redirect("admin_login")
   
    offer_type = request.GET.get('type', 'CATEGORY')
    
    # Validate offer type
    if offer_type not in ['CATEGORY', 'PRODUCT']:
        offer_type = 'CATEGORY'
    
    current_date = timezone.now().date()

    if offer_type == 'CATEGORY':
        offers = Offer.objects.filter(scope='CATEGORY').select_related('category')
    else:
        offers = Offer.objects.filter(scope='PRODUCT').select_related('product')

    # Update status of expired offers
    for offer in offers:
        # Convert offer.end_date to date if it's a datetime
        offer_end_date = offer.end_date.date() if hasattr(offer.end_date, 'date') else offer.end_date
        
        if offer_end_date < current_date and offer.status:
            offer.status = False
            offer.save()

    # Add pagination
    paginator = Paginator(offers, 2)  # Show 10 offers per page
    page = request.GET.get('page')
    
    try:
        offers_page = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        offers_page = paginator.page(1)
    except EmptyPage:
        # If page is out of range, deliver last page of results
        offers_page = paginator.page(paginator.num_pages)
    
    context = {
        'offers': offers,
        'offer_type': offer_type,
        'paginator': paginator,
    }
    return render(request, 'offer.html', context)

#this view for edditing  offer 
def edit_offer(request, offer_id):
    if not request.user.is_staff:
        return redirect("admin_login")

    # Get the offer object or return 404 if not found
    offer = get_object_or_404(Offer, id=offer_id)
    print(f"Offer ID: {offer.id}")

    if request.method == "POST":
        # Update the offer with form data
        offer.name = request.POST.get('name')
        offer.scope = request.POST.get('scope')
        offer.offer_type = request.POST.get('offer_type')
        offer.discount = request.POST.get('discount')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        offer.start_date = datetime.strptime(start_date_str, '%Y-%m-%dT%H:%M') if start_date_str else None
        offer.end_date = datetime.strptime(end_date_str, '%Y-%m-%dT%H:%M') if end_date_str else None
        offer.status = request.POST.get('status') == 'on'
        offer.category = Category.objects.get(id=request.POST.get('category')) if request.POST.get('category') else None
        offer.product = Product.objects.get(id=request.POST.get('product')) if request.POST.get('product') else None

        # Validate and save
        try:
            offer.clean()  # Run model validation
            offer.save()
            messages.success(request, "Offer updated successfully!")
            return redirect('offer_management')
        except Exception as e:
            messages.error(request, f"Error updating offer: {str(e)}")

    # For GET request, prepare the context
    categories = Category.objects.filter(is_deleted=False, status=True)
    products = Product.objects.filter(is_deleted=False, status=True)
    context = {
        'offer': offer,
        'categories': categories,
        'products': products,
        'form': {
            'name': offer.name,
            'scope': offer.scope,
            'category': offer.category,
            'product': offer.product,
            'offer_type': offer.offer_type,
            'discount': offer.discount,
            'start_date': (offer.start_date.strftime('%Y-%m-%dT%H:%M') if isinstance(offer.start_date, datetime) else offer.start_date) if offer.start_date else '',
            'end_date': (offer.end_date.strftime('%Y-%m-%dT%H:%M') if isinstance(offer.end_date, datetime) else offer.end_date) if offer.end_date else '',            'status': offer.status,
        }
    }
    return render(request, 'edit_offer.html', context)

#this view for adding  offer 
def add_offer(request):
    if not request.user.is_staff:
        return redirect("admin_login")
    
    if request.method == 'POST':
        form = OfferForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('offer_management')
    else:
        form = OfferForm()
        
    return render(request, 'add_offer.html', {'form': form})

#this view for deletting  offer 
@require_http_methods(["POST"])
def delete_offer(request, offer_id):
    if not request.user.is_staff:
        return redirect("admin_login")
 
    try:
        with transaction.atomic():
            offer = get_object_or_404(Offer, id=offer_id)
            offer.delete()
        
        return JsonResponse({
            'status': 'success', 
            'message': 'Offer deleted successfully'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error', 
            'message': str(e)
        }, status=400)
    
#this view for adding  offer 
def add_offer(request):
    if not request.user.is_staff:
        return redirect("admin_login")
    
    categories = Category.objects.all()
    products = Product.objects.all()

    if request.method == 'POST':
        form = OfferForm(request.POST)
        if form.is_valid():
            try:
                
                offer = form.save()
                messages.success(request, f"Offer '{offer.name}' created successfully!")
                return redirect('offer_management')  
            except Exception as e:
                messages.error(request, f"Error creating offer: {str(e)}")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = OfferForm()

    context = {
        'form': form,
        'categories': categories,
        'products': products,
    }
    return render(request, 'add_offer.html', context)

def is_admin(user):
    return user.is_superuser or user.is_staff

#this view for seeing  coupon 
def coupon_list(request):
    if not request.user.is_staff:
        return redirect("admin_login")
    
    search_query = request.GET.get('search', '')
    
    if search_query:
        coupons = Coupon.objects.filter(
            Q(code__icontains=search_query) | 
            Q(description__icontains=search_query)
        ).order_by('-created_at')
    else:
        coupons = Coupon.objects.all().order_by('-created_at')
    
    page = request.GET.get('page', 1)
    paginator = Paginator(coupons, 2)  # Show 10 coupons per page
    
    try:
        coupons = paginator.page(page)
    except PageNotAnInteger:
        coupons = paginator.page(1)
    except EmptyPage:
        coupons = paginator.page(paginator.num_pages)
    
    context = {
        'coupons': coupons,
        'search_query': search_query,
    }
    
    return render(request, 'coupon_list.html', context)

#this view for adding  coupon 
def add_coupon(request):
    if not request.user.is_staff:
        return redirect("admin_login")
    
    if request.method == 'POST':
        code = request.POST.get('code')
        description = request.POST.get('description')
        discount_type = request.POST.get('discount_type')
        discount_value = request.POST.get('discount_value')
        min_purchase_amount = request.POST.get('min_purchase_amount')
        max_discount_amount = request.POST.get('max_discount_amount') or None
        valid_from = request.POST.get('valid_from')
        valid_until = request.POST.get('valid_until')
        usage_limit = request.POST.get('usage_limit') or None
        per_user_limit = request.POST.get('per_user_limit')
        is_active = 'is_active' in request.POST
        
        if not code:
            messages.error(request, "Coupon code is required")
            return render(request, 'add_coupon.html')
        
        if Coupon.objects.filter(code=code).exists():
            messages.error(request, "Coupon code already exists")
            return render(request, 'add_coupon.html')
        
        try:
            coupon = Coupon(
                code=code,
                description=description,
                discount_type=discount_type,
                discount_value=discount_value,
                min_purchase_amount=min_purchase_amount,
                max_discount_amount=max_discount_amount,
                valid_from=valid_from,
                valid_until=valid_until,
                usage_limit=usage_limit,
                per_user_limit=per_user_limit,
                is_active=is_active
            )
            coupon.save()
            messages.success(request, "Coupon added successfully")
            return redirect('coupon_list')
        except Exception as e:
            messages.error(request, f"Error adding coupon: {str(e)}")
            return render(request, 'add_coupon.html')
    
    return render(request, 'add_coupon.html')

#this view for edditing  coupon
def edit_coupon(request, coupon_id):
    if not request.user.is_staff:
        return redirect("admin_login")
    coupon = get_object_or_404(Coupon, id=coupon_id)
    
    if request.method == 'POST':
        coupon.code = request.POST.get('code')
        coupon.description = request.POST.get('description')
        coupon.discount_type = request.POST.get('discount_type')
        coupon.discount_value = request.POST.get('discount_value')
        coupon.min_purchase_amount = request.POST.get('min_purchase_amount')
        
        max_discount = request.POST.get('max_discount_amount')
        coupon.max_discount_amount = max_discount if max_discount else None
        
        coupon.valid_from = request.POST.get('valid_from')
        coupon.valid_until = request.POST.get('valid_until')
        
        usage_limit = request.POST.get('usage_limit')
        coupon.usage_limit = usage_limit if usage_limit else None
        
        coupon.per_user_limit = request.POST.get('per_user_limit')
        coupon.is_active = 'is_active' in request.POST
        
        if Coupon.objects.filter(code=coupon.code).exclude(id=coupon_id).exists():
            messages.error(request, "Coupon code already exists")
            return render(request, 'edit_coupon.html', {'coupon': coupon})
        
        try:
            coupon.save()
            messages.success(request, "Coupon updated successfully")
            return redirect('coupon_list')
        except Exception as e:
            messages.error(request, f"Error updating coupon: {str(e)}")
    
    context = {
        'coupon': coupon
    }
    
    return render(request, 'edit_coupon.html', context)

#this view for deleting  coupon 
def delete_coupon(request, coupon_id):
    if not request.user.is_staff:
        return redirect("admin_login")      
    coupon = get_object_or_404(Coupon, id=coupon_id)
    
    try:
        if coupon.usage_count > 0:
            coupon.is_active = False
            coupon.save()
            messages.success(request, f"Coupon '{coupon.code}' has been deactivated since it has already been used")
        else:
            coupon.delete()
            messages.success(request, f"Coupon '{coupon.code}' deleted successfully")
    except Exception as e:
        messages.error(request, f"Error deleting coupon: {str(e)}")
    
    return redirect('coupon_list')

def sales_report(request):
    # Default filter: Last 7 days
    end_date = timezone.now()
    start_date = end_date - timedelta(days=7)
    filter_type = request.GET.get('filter_type', 'custom')
    
    # Handle filter type
    if filter_type == 'daily':
        start_date = end_date - timedelta(days=0)
    elif filter_type == 'weekly':
        start_date = end_date - timedelta(days=6)  # 7 days including today
    elif filter_type == 'monthly':
        start_date = end_date - timedelta(days=30)
    elif filter_type == 'yearly':
        start_date = end_date - timedelta(days=365)
    else:  # custom
        try:
            start_date = timezone.datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d')
            end_date = timezone.datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d')
            if not timezone.is_aware(start_date):
                start_date = timezone.make_aware(start_date)
            if not timezone.is_aware(end_date):
                end_date = timezone.make_aware(end_date)
        except (ValueError, TypeError):
            pass  # Use default if invalid

    # Ensure timezone awareness
    start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
    end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)

    # Fetch orders within the date range
    orders = Order.objects.filter(
        created_at__range=(start_date, end_date),
        status__in=['Delivered']  # Adjust statuses as needed
    ).select_related('user', 'shipping_address').prefetch_related('items__product', 'items__variant')

    # Prepare order data with required details
    order_data = []
    for order in orders:
        products = ", ".join(
            f"{item.product.name} ({item.quantity})" + 
            (f" - {item.variant.ram}GB/{item.variant.storage}GB/{item.variant.color}" if item.variant else "")
            for item in order.items.all()
        )
        order_data.append({
            'order_id': order.id,
            'date': order.created_at,
            'customer': order.user.username,
            'products': products or "No items",
            'payment': order.get_payment_method_display(),
            'status': order.get_status_display(),
            'amount': order.calculate_total_price(),  # Includes shipping fee
        })

    # Prepare context
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'filter_type': filter_type,
        'order_data': order_data,
    }

    # Handle download requests (you'll need to implement these functions)
    download_format = request.GET.get('download', None)
    if download_format == 'pdf':
        return generate_pdf_report(context)
    elif download_format == 'excel':
        return generate_excel_report(context)

    return render(request, 'sales_report.html', context)

def generate_pdf_report(context):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                            rightMargin=40, leftMargin=40,  # Increased margins for better whitespace
                            topMargin=40, bottomMargin=40)
    elements = []

    # Enhanced styles
    styles = getSampleStyleSheet()
    
    # Brand header style
    brand_style = ParagraphStyle(
        'Brand',
        parent=styles['Title'],
        fontName='Helvetica-Bold',
        fontSize=24,
        alignment=1,  # Center
        spaceAfter=6,
        textColor=colors.HexColor("#1A5276")  # Dark blue for brand identity
    )
    
    # Title style
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Title'],
        fontName='Helvetica-Bold',
        fontSize=18,
        alignment=1,  # Center
        spaceAfter=10,
        textColor=colors.HexColor("#2E86C1")  # Blue for titles
    )
    
    # Subtitle style
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        alignment=1,  # Center
        spaceAfter=20,
        textColor=colors.HexColor("#5D6D7E")  # Slate gray for subtitles
    )
    
    # Normal text style
    normal_style = ParagraphStyle(
        'Normal',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#333333")  # Dark gray for normal text
    )
    
    # Summary style
    summary_style = ParagraphStyle(
        'Summary',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11,
        alignment=2,  # Right aligned
        textColor=colors.HexColor("#2E86C1")  # Blue for summary
    )

    # Add company logo/name at the top (replace with your company name)
    elements.append(Paragraph("Phonix", brand_style))
    
    # Title with improved styling
    elements.append(Paragraph("SALES REPORT", title_style))
    
    # Period and generation info with better styling - Indian format (DD-MM-YYYY)
    period_text = f"Period: {context['start_date'].strftime('%d-%m-%Y')} - {context['end_date'].strftime('%d-%m-%Y')}"
    elements.append(Paragraph(period_text, subtitle_style))
    
    # Use Django's timezone.localtime() for India time (already set in settings.py with TIME_ZONE = 'Asia/Kolkata')
    current_time = timezone.localtime(timezone.now())
    generated_text = "Generated on: " + current_time.strftime('%d-%m-%Y %I:%M %p')
    elements.append(Paragraph(generated_text, subtitle_style))

    # Space before table
    elements.append(Paragraph("<br/>", normal_style))

    # Table Headers - with better alignment and more descriptive headers
    data = [['Order #', 'Date', 'Customer', 'Product Details', 'Payment Method', 'Status', 'Amount (₹)']]
    
    # Cell content style for wrapping text
    wrap_style = ParagraphStyle(
        'wrap', 
        fontName='Helvetica',
        fontSize=9,
        leading=11
    )
    
    # Process order data with proper formatting
    total_amount = 0
    for order in context['order_data']:
        # Format product text as a paragraph to allow wrapping
        product_text = Paragraph(order['products'], wrap_style)
        
        # Convert order date to local time if it's timezone-aware
        order_date = order['date']
        if timezone.is_aware(order_date):
            order_date = timezone.localtime(order_date)
        
        # Add row to table with improved formatting - Indian date format
        data.append([
            str(order['order_id']),
            order_date.strftime('%d-%m-%Y %I:%M %p'),  # Date in DD-MM-YYYY HH:MM AM/PM format
            Paragraph(order['customer'], wrap_style),
            product_text,
            order['payment'],
            order['status'],
            f"{order['amount']:,.2f}",  # Comma as thousand separator
        ])
        
        total_amount += order['amount']

    # Table configuration with column widths optimized for content - adjusted for longer date format
    table = Table(data, repeatRows=1, colWidths=[45, 85, 80, 130, 75, 60, 65])
    
    # Zebra striping for better readability (alternate row colors)
    row_colors = [colors.HexColor("#F5F8FA"), colors.HexColor("#EBF5FB")]
    
    # Enhanced table styling
    table_style = TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1A5276")),  # Dark blue header
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        # Data cell styling
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        
        # Alignment by column type
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Order ID centered
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Date centered
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),    # Customer left-aligned
        ('ALIGN', (3, 1), (3, -1), 'LEFT'),    # Products left-aligned
        ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # Payment centered
        ('ALIGN', (5, 1), (5, -1), 'CENTER'),  # Status centered
        ('ALIGN', (6, 1), (6, -1), 'RIGHT'),   # Amount right-aligned
        
        # Grid styling
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#D5D8DC")),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Header bottom border more prominent
        ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor("#154360")),
    ])
    
    # Add zebra striping (alternate row colors)
    for i in range(1, len(data)):
        bg_color = row_colors[(i - 1) % len(row_colors)]
        table_style.add('BACKGROUND', (0, i), (-1, i), bg_color)
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Add total amount summary
    elements.append(Paragraph("<br/>", normal_style))
    elements.append(Paragraph(f"Total Sales Amount: ₹{total_amount:,.2f}", summary_style))
    
    # Add footer with additional information
    elements.append(Paragraph("<br/><br/>", normal_style))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        alignment=1,  # Center
        textColor=colors.HexColor("#7F8C8D")  # Light gray for footer
    )
    elements.append(Paragraph("This is an official sales report document. For inquiries, contact thomasbaby273@gmail.com", footer_style))

    # Build the PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

def generate_excel_report(context):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'

    # Create a workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sales Report"
    
    # Define colors for branding consistency with PDF report
    header_fill = openpyxl.styles.PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")  # Dark Blue
    alternate_row_fill = openpyxl.styles.PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")  # Light Blue
    title_font = openpyxl.styles.Font(name='Calibri', size=14, bold=True, color="2E86C1")  # Blue
    header_font = openpyxl.styles.Font(name='Calibri', size=11, bold=True, color="FFFFFF")  # White
    normal_font = openpyxl.styles.Font(name='Calibri', size=10)
    amount_font = openpyxl.styles.Font(name='Calibri', size=10, bold=True)
    
    # Create title row
    worksheet.merge_cells('A1:G1')
    worksheet.merge_cells('A2:G2')
    worksheet.merge_cells('A3:G3')
    
    # Company name
    cell = worksheet.cell(row=1, column=1, value="Phonix")
    cell.font = openpyxl.styles.Font(name='Calibri', size=16, bold=True, color="1A5276")
    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    # Report title
    cell = worksheet.cell(row=2, column=1, value="SALES REPORT")
    cell.font = title_font
    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    # Report period
    period_text = f"Period: {context['start_date'].strftime('%d-%m-%Y')} - {context['end_date'].strftime('%d-%m-%Y')}"
    cell = worksheet.cell(row=3, column=1, value=period_text)
    cell.font = openpyxl.styles.Font(name='Calibri', size=10, color="5D6D7E")  # Slate gray
    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    # Add space before headers
    current_row = 5
    
    # Add headers
    headers = ['Order #', 'Date', 'Customer', 'Product Details', 'Payment Method', 'Status', 'Amount (₹)']
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=current_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        # Apply border to headers
        cell.border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin', color='D5D8DC'),
            right=openpyxl.styles.Side(style='thin', color='D5D8DC'),
            top=openpyxl.styles.Side(style='thin', color='D5D8DC'),
            bottom=openpyxl.styles.Side(style='medium', color='154360')
        )
    
    # Add data rows
    current_row += 1
    row_num = 0
    total_amount = 0
    
    for order in context['order_data']:
        row_num += 1
        
        # Convert order date to local time if it's timezone-aware
        order_date = order['date']
        if timezone.is_aware(order_date):
            order_date = timezone.localtime(order_date)
            
        # Format date as string for Excel
        date_str = order_date.strftime('%d-%m-%Y %I:%M %p')  # DD-MM-YYYY HH:MM AM/PM
        
        # Add row data
        row_data = [
            order['order_id'],
            date_str,
            order['customer'],
            order['products'],
            order['payment'],
            order['status'],
            float(order['amount'])
        ]
        
        # Apply zebra striping
        row_fill = alternate_row_fill if row_num % 2 == 0 else None
        
        # Write data to cells with appropriate formatting
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=current_row, column=col_num, value=value)
            cell.font = normal_font
            
            # Apply cell alignment based on content type
            if col_num == 1:  # Order ID
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            elif col_num == 2:  # Date
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            elif col_num in (3, 4):  # Customer, Products
                cell.alignment = openpyxl.styles.Alignment(horizontal='left', wrap_text=True)
            elif col_num in (5, 6):  # Payment, Status
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            elif col_num == 7:  # Amount
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                cell.number_format = '₹#,##0.00'
                cell.font = amount_font
            
            # Apply fill and border
            if row_fill:
                cell.fill = row_fill
                
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin', color='D5D8DC'),
                right=openpyxl.styles.Side(style='thin', color='D5D8DC'),
                top=openpyxl.styles.Side(style='thin', color='D5D8DC'),
                bottom=openpyxl.styles.Side(style='thin', color='D5D8DC')
            )
        
        total_amount += order['amount']
        current_row += 1
    
    # Add total row
    current_row += 1
    worksheet.merge_cells(f'A{current_row}:F{current_row}')
    
    total_label = worksheet.cell(row=current_row, column=1, value="Total Sales Amount:")
    total_label.font = openpyxl.styles.Font(name='Calibri', size=11, bold=True, color="2E86C1")
    total_label.alignment = openpyxl.styles.Alignment(horizontal='right')
    
    total_value = worksheet.cell(row=current_row, column=7, value=float(total_amount))
    total_value.font = openpyxl.styles.Font(name='Calibri', size=11, bold=True, color="2E86C1")
    total_value.alignment = openpyxl.styles.Alignment(horizontal='right')
    total_value.number_format = '₹#,##0.00'
    
    # Add footer
    current_row += 2
    worksheet.merge_cells(f'A{current_row}:G{current_row}')
    footer = worksheet.cell(row=current_row, column=1, 
                           value="This is an official sales report document. For inquiries, contact thomasbaby273@gmail.com")
    footer.font = openpyxl.styles.Font(name='Calibri', size=8, color="7F8C8D")  # Light gray
    footer.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    # Set column widths
    column_widths = {
        1: 10,  # Order #
        2: 18,  # Date
        3: 20,  # Customer
        4: 40,  # Product Details
        5: 15,  # Payment Method
        6: 12,  # Status
        7: 15,  # Amount
    }
    
    for col_num, width in column_widths.items():
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
    # Auto-filter headers to make report interactive
    worksheet.auto_filter.ref = f"A{5}:G{5}"
    
    # Save the workbook to the response
    workbook.save(response)
    return response
    
#this view for admin-logout
def admin_logout(request):
    logout(request)
    return redirect("admin_login")
